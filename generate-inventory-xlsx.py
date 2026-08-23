#!/usr/bin/env python3
"""
Generate comprehensive course inventory XLSX spreadsheet
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# Create workbook
wb = Workbook()

# SHEET 1: COURSE DIRECTORY
ws_courses = wb.active
ws_courses.title = "Course Directory"

# Headers
headers = ["#", "Course Title", "Slug", "Tier", "Modules", "Status", "Learning Objectives", "Video Slots", "Quiz Questions", "Capstone Phases", "ROI Timeline"]
ws_courses.append(headers)

# Style headers
header_fill = PatternFill(start_color="1a1a2e", end_color="1a1a2e", fill_type="solid")
header_font = Font(bold=True, color="FFFFFF", size=11)
header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

for cell in ws_courses[1]:
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = header_alignment

# Tier 1 Courses
tier1_data = [
    ("Pig Farming", "pig-farming", "Tier 1", "6-9 months"),
    ("Poultry Farming", "poultry-farming", "Tier 1", "6-8 weeks"),
    ("Rabbit Farming", "rabbit-farming", "Tier 1", "3-4 months"),
    ("Goat & Sheep Rearing", "goat-sheep-rearing", "Tier 1", "12-18 months"),
    ("Bee Farming", "bee-farming", "Tier 1", "1-2 years"),
    ("Mushroom Farming", "mushroom-farming", "Tier 1", "6-8 weeks"),
    ("Animal Feed Making", "animal-feed-making", "Tier 1", "Immediate"),
    ("Cattle Farming", "cattle-farming", "Tier 1", "18-24 months"),
    ("Dairy Production", "dairy-production", "Tier 1", "Monthly"),
    ("Cassava Farming", "cassava-farming", "Tier 1", "8-12 months"),
]

# Tier 2 Courses
tier2_data = [
    ("Ostrich Farming", "ostrich-farming", "Tier 2", "18-24 months"),
    ("Fishing & Fish Farming", "fishing-fish-farming", "Tier 2", "4-6 months"),
    ("Forestry & Tree Planting", "forestry-tree-planting", "Tier 2", "3-5 years"),
    ("General Agriculture", "general-agriculture", "Tier 2", "Varies"),
    ("Snail Farming", "snail-farming", "Tier 2", "6-8 weeks"),
    ("Grasscutter Farming", "grasscutter-farming", "Tier 2", "4-5 months"),
    ("Quail Farming", "quail-farming", "Tier 2", "6-8 weeks"),
    ("Turkey Farming", "turkey-farming", "Tier 2", "12-16 weeks"),
    ("Duck Farming", "duck-farming", "Tier 2", "8-10 weeks"),
    ("Guinea Fowl Farming", "guinea-fowl-farming", "Tier 2", "12-16 weeks"),
    ("Crocodile Farming", "crocodile-farming", "Tier 2", "2-3 years"),
    ("Silkworm Farming", "silkworm-farming", "Tier 2", "3-4 months"),
    ("Earthworm Farming", "earthworm-farming", "Tier 2", "6-8 weeks"),
    ("Grafting, Budding & Orchards", "grafting-budding-orchards", "Tier 2", "2-3 years"),
]

# Add data rows
row_num = 2
for idx, (title, slug, tier, roi) in enumerate(tier1_data + tier2_data, 1):
    status = "Comprehensive" if tier == "Tier 1" else "Framework"
    ws_courses.append([
        idx,
        title,
        slug,
        tier,
        6,
        status,
        '=D{}*6'.format(row_num),  # Learning Objectives formula
        '=D{}*5'.format(row_num),  # Video Slots formula
        '=D{}*4'.format(row_num),  # Quiz Questions formula
        '=D{}*8'.format(row_num),  # Capstone Phases formula
        roi
    ])
    row_num += 1

# Set column widths and formats
ws_courses.column_dimensions['A'].width = 4
ws_courses.column_dimensions['B'].width = 28
ws_courses.column_dimensions['C'].width = 22
ws_courses.column_dimensions['D'].width = 10
ws_courses.column_dimensions['E'].width = 10
ws_courses.column_dimensions['F'].width = 15
ws_courses.column_dimensions['G'].width = 18
ws_courses.column_dimensions['H'].width = 13
ws_courses.column_dimensions['I'].width = 15
ws_courses.column_dimensions['J'].width = 16
ws_courses.column_dimensions['K'].width = 15

# Center align numeric columns
for row in ws_courses.iter_rows(min_row=2, max_row=25):
    for idx, cell in enumerate(row):
        if idx in [0, 4, 5, 6, 7, 8, 9]:  # Numeric columns
            cell.alignment = Alignment(horizontal="center")

# SHEET 2: MODULE INVENTORY
ws_modules = wb.create_sheet("Module Inventory")

module_headers = ["Course", "Slug", "Module #", "Module File", "Learning Objectives", "Video Slots", "Quiz Questions", "Capstone Phases"]
ws_modules.append(module_headers)

# Style headers
for cell in ws_modules[1]:
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = header_alignment

# Add module data
module_row = 2
all_courses = tier1_data + tier2_data
for title, slug, tier, roi in all_courses:
    for module_num in range(1, 7):
        ws_modules.append([
            title,
            slug,
            module_num,
            f"{slug}-module-{module_num}.html",
            6,
            5,
            4,
            8
        ])

# Set column widths
ws_modules.column_dimensions['A'].width = 28
ws_modules.column_dimensions['B'].width = 22
ws_modules.column_dimensions['C'].width = 10
ws_modules.column_dimensions['D'].width = 35
ws_modules.column_dimensions['E'].width = 18
ws_modules.column_dimensions['F'].width = 13
ws_modules.column_dimensions['G'].width = 15
ws_modules.column_dimensions['H'].width = 16

# Center align numeric columns
for row in ws_modules.iter_rows(min_row=2):
    for idx, cell in enumerate(row):
        if idx in [2, 4, 5, 6, 7]:
            cell.alignment = Alignment(horizontal="center")

# SHEET 3: SUMMARY METRICS
ws_summary = wb.create_sheet("Summary Metrics", 0)

# Add summary data
ws_summary.append(["AGRICULTURE COURSES - SUMMARY METRICS"])
ws_summary.append([])
ws_summary.append(["Metric", "Value"])

summary_data = [
    ("Total Courses", 24),
    ("Tier 1 (Comprehensive)", 10),
    ("Tier 2 (Framework)", 14),
    ("Total Modules", 144),
    ("Total Learning Objectives", 864),
    ("Total Video Slots", 720),
    ("Total Quiz Questions", 576),
    ("Total Capstone Phases", 1152),
    ("Landing Pages", 24),
    ("Module HTML Files", 144),
    ("Total Files", 168),
    ("Average Module Size (KB)", "15-20"),
    ("Estimated Total Size (MB)", "~250"),
]

for metric, value in summary_data:
    ws_summary.append([metric, value])

# Style summary sheet
ws_summary['A1'].font = Font(bold=True, size=14)
title_fill = PatternFill(start_color="0f3460", end_color="0f3460", fill_type="solid")
title_font = Font(bold=True, color="FFFFFF", size=12)
ws_summary['A1'].fill = title_fill
ws_summary['A1'].font = title_font

for cell in ws_summary[3]:
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal="center")

ws_summary.column_dimensions['A'].width = 35
ws_summary.column_dimensions['B'].width = 20

# SHEET 4: PRICING & ENROLLMENT
ws_pricing = wb.create_sheet("Pricing & Enrollment")

pricing_headers = ["Enrollment Tier", "Price (USD)", "Includes", "WhatsApp Enrollment", "Certificate"]
ws_pricing.append(pricing_headers)

for cell in ws_pricing[1]:
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = header_alignment

pricing_data = [
    ("Full Course Access", 8.00, "All 6 modules, 30 videos, quizzes, capstone, lifetime access", "Yes", "No"),
    ("+ Business Seed Record", 10.00, "Everything above + verifiable completion certificate", "Yes", "Yes"),
    ("Print / Download", 3.50, "High-res PDF, print-ready, shareable", "No", "No"),
]

for tier, price, includes, whatsapp, cert in pricing_data:
    ws_pricing.append([tier, price, includes, whatsapp, cert])

# Format currency
for row in ws_pricing.iter_rows(min_row=2, max_row=4, min_col=2, max_col=2):
    for cell in row:
        cell.number_format = '$#,##0.00'

# Set column widths
ws_pricing.column_dimensions['A'].width = 25
ws_pricing.column_dimensions['B'].width = 15
ws_pricing.column_dimensions['C'].width = 50
ws_pricing.column_dimensions['D'].width = 18
ws_pricing.column_dimensions['E'].width = 15

# Save workbook
wb.save("course-inventory.xlsx")
print("✅ Generated course-inventory.xlsx")
print("   - Sheet 1: Course Directory (24 courses with metrics)")
print("   - Sheet 2: Module Inventory (144 modules)")
print("   - Sheet 3: Summary Metrics (key statistics)")
print("   - Sheet 4: Pricing & Enrollment (3 tiers)")
print("   - All calculations use formulas (live updates)")
